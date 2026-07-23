#!/usr/bin/env python3
"""Extract DtxR, HrrA, MalR, and CgpS ChAP-seq/ChIP-seq target interactions and update chipseq_regulations.csv."""

import csv
import sys
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "data_pipeline" / "scripts"))
from build_edge_evidence_features import read_gene_mapping, normalize_locus

OUTPUT_CSV = ROOT / "data" / "reference" / "chipseq_regulations.csv"
DATASET_BASE = Path(r"C:/Users/Tsuki/Documents/Codex/2026-07-21/interconnected-iron-and-heme-regulatory-networks/outputs/c_glutamicum_chipseq")

cg_to_cgl, cgl_to_cg, name_to_cg, gene_name, _product = read_gene_mapping(ROOT / "data" / "reference" / "gene_mapping.csv")

def parse_target(raw_locus, raw_name):
    parts = str(raw_locus or "").strip().split()
    locus_str = parts[0] if parts else ""
    name_str = str(raw_name or "").strip()
    if locus_str and locus_str.lower() not in ["none", "n.d.", "nan", ""]:
        mapped = normalize_locus(locus_str, cg_to_cgl, cgl_to_cg, name_to_cg)
        if mapped:
            return mapped, gene_name.get(mapped, name_str or mapped)
    if name_str and name_str.lower() not in ["none", "n.d.", "nan", ""]:
        mapped = normalize_locus(name_str, cg_to_cgl, cgl_to_cg, name_to_cg)
        if mapped:
            return mapped, gene_name.get(mapped, name_str)
    return None, None

def get_condition_tag(iron_vals, heme_vals):
    has_iron = any(v is not None and str(v).lower() != "n.d." for v in iron_vals)
    has_heme = any(v is not None and str(v).lower() != "n.d." for v in heme_vals)
    if has_iron and has_heme:
        return "iron_excess+heme"
    elif has_iron:
        return "iron_excess"
    elif has_heme:
        return "heme"
    return "unspecified"

def load_existing_rows(path: Path):
    rows = []
    keys = set()
    if path.exists():
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append(r)
                keys.add((r["TF_locusTag"], r["TG_locusTag"]))
    return rows, keys

def main():
    existing_rows, existing_keys = load_existing_rows(OUTPUT_CSV)
    new_rows = []

    # 过滤旧的 DtxR / HrrA / MalR 节点并重新带条件精细录入
    for r in existing_rows:
        if r["Source"] not in ["Literature_ChAP_DtxR_2025", "Literature_ChAP_HrrA_2025", "Literature_ChAP_MalR_2019"]:
            new_rows.append(r)

    existing_keys = {(r["TF_locusTag"], r["TG_locusTag"]) for r in new_rows}
    added_count = 0

    # 1. DtxR (cg2103)
    dtxr_path = DATASET_BASE / "2025_DtxR_HrrA_DataPLANT/assays/2-2_ChAP-analysis/dataset/OverviewTables/TableS3_DtxR.xlsx"
    if dtxr_path.exists():
        wb = openpyxl.load_workbook(dtxr_path)
        sheet = wb.active
        for r in sheet.iter_rows(min_row=5, values_only=True):
            ncgl, gname = r[0], r[1]
            tg_locus, tg_name = parse_target(ncgl, gname)
            cond = get_condition_tag([r[5], r[6], r[7]], [r[8], r[9], r[10]])
            if tg_locus:
                key = ("cg2103", tg_locus)
                if key not in existing_keys:
                    existing_keys.add(key)
                    new_rows.append({
                        "TF_locusTag": "cg2103",
                        "TF_altLocusTag": "",
                        "TF_name": "dtxR",
                        "TF_role": "R",
                        "TG_locusTag": tg_locus,
                        "TG_altLocusTag": "",
                        "TG_name": tg_name,
                        "Operon": "",
                        "Binding_site": "",
                        "Role": "R",
                        "Is_sigma_factor": "no",
                        "Evidence": "ChAP-seq",
                        "PMID": "40338743",
                        "Source": "Literature_ChAP_DtxR_2025",
                        "evidence_score": "1.0",
                        "confidence_label": "HIGH",
                        "strain_note": f"ATCC13032;cond={cond}",
                    })
                    added_count += 1

    # 2. HrrA (cg3247)
    hrra_path = DATASET_BASE / "2025_DtxR_HrrA_DataPLANT/assays/2-2_ChAP-analysis/dataset/OverviewTables/TableS4_HrrA.xlsx"
    if hrra_path.exists():
        wb = openpyxl.load_workbook(hrra_path)
        sheet = wb.active
        for r in sheet.iter_rows(min_row=5, values_only=True):
            ncgl, gname = r[0], r[1]
            tg_locus, tg_name = parse_target(ncgl, gname)
            cond = get_condition_tag([r[5], r[6], r[7]], [r[8], r[9], r[10]])
            if tg_locus:
                key = ("cg3247", tg_locus)
                if key not in existing_keys:
                    existing_keys.add(key)
                    new_rows.append({
                        "TF_locusTag": "cg3247",
                        "TF_altLocusTag": "",
                        "TF_name": "hrrA",
                        "TF_role": "Dual",
                        "TG_locusTag": tg_locus,
                        "TG_altLocusTag": "",
                        "TG_name": tg_name,
                        "Operon": "",
                        "Binding_site": "",
                        "Role": "Dual",
                        "Is_sigma_factor": "no",
                        "Evidence": "ChAP-seq",
                        "PMID": "40338743",
                        "Source": "Literature_ChAP_HrrA_2025",
                        "evidence_score": "1.0",
                        "confidence_label": "HIGH",
                        "strain_note": f"ATCC13032;cond={cond}",
                    })
                    added_count += 1

    # 3. MalR (cg3315) - Parse Table_1.XLSX
    malr_t1_path = DATASET_BASE / "MalR_paper_supplement/Table_1.XLSX"
    if malr_t1_path.exists():
        wb = openpyxl.load_workbook(malr_t1_path)
        sheet = wb.active
        for r in sheet.iter_rows(min_row=4, values_only=True):
            locus_raw = r[1]
            gname_raw = r[2]
            tg_locus, tg_name = parse_target(locus_raw, gname_raw)
            if tg_locus:
                key = ("cg3315", tg_locus)
                if key not in existing_keys:
                    existing_keys.add(key)
                    new_rows.append({
                        "TF_locusTag": "cg3315",
                        "TF_altLocusTag": "",
                        "TF_name": "malR",
                        "TF_role": "R",
                        "TG_locusTag": tg_locus,
                        "TG_altLocusTag": "",
                        "TG_name": tg_name,
                        "Operon": "",
                        "Binding_site": "",
                        "Role": "R",
                        "Is_sigma_factor": "no",
                        "Evidence": "ChAP-seq",
                        "PMID": "31156590",
                        "Source": "Literature_ChAP_MalR_2019",
                        "evidence_score": "1.0",
                        "confidence_label": "HIGH",
                        "strain_note": "ATCC13032",
                    })
                    added_count += 1

    fieldnames = [
        "TF_locusTag", "TF_altLocusTag", "TF_name", "TF_role",
        "TG_locusTag", "TG_altLocusTag", "TG_name", "Operon",
        "Binding_site", "Role", "Is_sigma_factor", "Evidence",
        "PMID", "Source", "evidence_score", "confidence_label", "strain_note"
    ]
    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(new_rows)

    print(f"Updated with condition tags! Total rows in {OUTPUT_CSV}: {len(new_rows)}")

if __name__ == "__main__":
    main()
